import csv
import os
from re import match
from time import sleep
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.formatting.rule import FormulaRule
from TransferHero import assistDotOrg
from TransferHero.search import find_class_schedule_link
from TransferHero.course import Course

APP_NAME = "TransferHero"

HAMMER_DELAY = 3
DOWNLOAD_FOLDER = r'S:\assist'
INPUT_FILENAME = r'comp132.csv'
SCHOOL_NAME = 'CSUMB'
MAJOR_NAME = 'Computer Science'
CATALOG_YEAR = '2020-2021'
TARGET_COURSE = 'CST 238'


if __name__ == '__main__':
    # Open the input csv and save the info we care about to a list of Course objects
    courses = []
    with open(INPUT_FILENAME, 'r') as csv_file:
        reader = csv.DictReader(csv_file)
        for row in reader:
            courses.append(Course(row['Institution'], row['Local Dept. Name & Number'], row['Local Course Title(s)']))

    # Download agreement PDFs for every community college in our list
    # print('Downloading course agreements...')
    # # Create the Selenium web driver we'll need for automation
    # driver = assistDotOrg.create_driver(download_folder=DOWNLOAD_FOLDER)
    # for course in sorted(courses):
    #     print(f'   {course}')
    #     assistDotOrg.get_agreement_pdf(driver, CATALOG_YEAR, SCHOOL_NAME, course.school, MAJOR_NAME)
    #     sleep(HAMMER_DELAY)  # Avoid hammering assist.org
    # driver.close()

    # Create a spreadsheet that we can store data in as we parse the PDFs we just downloaded
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f'{TARGET_COURSE} agreements'
    sheet.append(['School', 'Class', 'Description', 'Offered?', 'Notes'])  # Header row

    # Parse out our PDFs to verify which classes satisfy the class we need
    for filename in os.listdir(DOWNLOAD_FOLDER):
        absolute_filename = os.path.join(DOWNLOAD_FOLDER, filename)
        file_ext = os.path.splitext(filename)

        # Delete duplicate PDFs (i.e. identical catalogs from the same school)
        if match(r'\(\d{1,2}\)', file_ext[0][-3:]):
            os.remove(absolute_filename)
            continue

        # Parse the PDF and look for our target course
        if file_ext[1].lower() == '.pdf':
            school = assistDotOrg.parse_agreement_pdf(absolute_filename, TARGET_COURSE)
            if school:  # Target course was found in this school's agreement so the name was returned
                # Since we can't accurately parse the course name from the PDF, try to look it up using the list we
                #  created from our csv file
                this_course = Course(school, 'Unknown', 'Unknown')
                for course in courses:
                    if school == course.school:
                        this_course = course

                # Add the class to the spreadsheet
                sheet.append([this_course.school, this_course.course_code, this_course.description])

                # Try to find and add a link to the school's class schedule
                schedule_link = find_class_schedule_link(school)
                cell = sheet.cell(row=sheet.max_row, column=1)
                cell.hyperlink = schedule_link
                cell.style = 'Hyperlink'

    # Make our worksheet a little prettier
    # Bold the first row
    for col in range(1, 6):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
    # Fix column widths
    column_widths = []
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))
    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = column_width + 3
    # Conditional formatting for "Offered?" column
    red = 'ffc7ce'
    green = 'c6efce'
    red_fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
    green_fill = PatternFill(start_color=green, end_color=green, fill_type='solid')
    sheet.conditional_formatting.add('D2:D1000', FormulaRule(formula=['NOT(ISERROR(SEARCH("N",D2)))'], stopIfTrue=True, fill=red_fill))
    sheet.conditional_formatting.add('D2:D1000', FormulaRule(formula=['NOT(ISERROR(SEARCH("Y",D2)))'], stopIfTrue=True, fill=green_fill))

    # Save the spreadsheet
    output_filename = os.path.join(DOWNLOAD_FOLDER, f'{SCHOOL_NAME} {TARGET_COURSE} agreements.xlsx')
    workbook.save(output_filename)
