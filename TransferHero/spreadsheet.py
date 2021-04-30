from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.formatting.rule import FormulaRule


class Spreadsheet:
    def __init__(self, sheet_title, header_row):
        self.workbook = self.initialize_workbook(sheet_title, header_row)

    @staticmethod
    def initialize_workbook(sheet_title, header_row):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        sheet.append(header_row)
        # Make the header row bold
        for col in range(1, len(header_row) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
        return workbook

    def add_row(self, row_data):
        sheet = self.workbook.active
        sheet.append(row_data)

    def add_hyperlink(self, cell_row, cell_col, url):
        sheet = self.workbook.active
        cell = sheet.cell(row=cell_row, column=cell_col)
        cell.hyperlink = url
        cell.style = 'Hyperlink'

    def prettify(self):
        sheet = self.workbook.active

        # Define some background fill styles that we'll use for conditional formatting
        red = 'ffc7ce'
        green = 'c6efce'
        red_fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
        green_fill = PatternFill(start_color=green, end_color=green, fill_type='solid')

        # Iterate over the first row doing the following:
        #  Set bold font for all cells in the first row
        #  Set static column sizes since our auto resize isn't useful for this data set
        #  Collect letters of the cells where we want to apply conditional formatting
        columns_to_format = []
        for i, cell in enumerate(sheet[1]):
            column_letter = get_column_letter(i + 1)
            cell.font = Font(bold=True)
            if cell.value == 'School':
                sheet.column_dimensions[column_letter].width = 33
            elif cell.value == 'Class':
                sheet.column_dimensions[column_letter].width = 18
            elif cell.value in ['Description', 'Notes']:
                sheet.column_dimensions[column_letter].width = 66
            elif cell.value in ['Offered', 'Online', 'Synchronous']:
                sheet.column_dimensions[column_letter].width = 5
                columns_to_format.append(column_letter)

        # Conditional formatting rules to set background red if cell contains 'n', green if cell contains 'y'
        for column_letter in columns_to_format:
            sheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1000',
                                             FormulaRule(formula=[f'NOT(ISERROR(SEARCH("N",{column_letter}2)))'],
                                                         stopIfTrue=True,
                                                         fill=red_fill))
            sheet.conditional_formatting.add(f'{column_letter}2:{column_letter}1000',
                                             FormulaRule(formula=[f'NOT(ISERROR(SEARCH("Y",{column_letter}2)))'],
                                                         stopIfTrue=True,
                                                         fill=green_fill))

    def save(self, output_filename):
        self.prettify()
        retry = ''
        while retry != 'a':
            try:
                self.workbook.save(output_filename)
            except PermissionError:
                retry = input('Error: Failed to save spreadsheet. Is the file already open?\n'
                              '  Please close Excel and press enter to continue, or type a to abort')
            else:
                print(f'Done. Saved results to:\n'
                      f'{output_filename}')
                break
